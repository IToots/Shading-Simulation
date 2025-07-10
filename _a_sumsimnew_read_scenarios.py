# -*- coding: utf-8 -*-
"""
Created on Mon Aug 26 14:20:19 2024

@author: ruiui
"""

import openpyxl
import warnings
import os
import pandas as pd
import re
import matplotlib.pyplot as plt
from collections import defaultdict

from lists import Vals, W, x, Irradiance


circ = 'TCT'

# circ = 'TCT_D'

# circ = 'SP'

# circ = 'SP_D'



# # all cells identical
# csv_directory = "C:/Users/ruiui/Desktop/solar plug and play old stuff/iteration data/z_3x3"
# circuit_irr_directory = "C:/Users/ruiui/Desktop/solar plug and play old stuff/iteration data/_Irr_teste/3x3"
# m = 'identical'

# modified cells corrected
csv_directory = "C:/Users/ruiui/Desktop/solar plug and play old stuff/iteration data/z_3x3/mod_corr"
circuit_irr_directory = "C:/Users/ruiui/Desktop/solar plug and play old stuff/iteration data/_Irr_teste/3x3/mod_corr"
m = 'mod'

base_directory = 'C:/Users/ruiui/Desktop/solar plug and play old stuff/sun simulator new/'
directory = f'{base_directory}{circ}/'

# -----------------------------------------------------------------------------
# -----------------------------------------------------------------------------
# -----------------------------------------------------------------------------
# LTS results
# -----------------------------------------------------------------------------
# -----------------------------------------------------------------------------
# -----------------------------------------------------------------------------

# Initialization
def initialize_data_structures():
    return (
        defaultdict(list),  # t1
        defaultdict(list),  # percentage_diffs_pos
        defaultdict(list),  # percentage_diffs_val_mpp
        defaultdict(list),  # percentage_diffs
        defaultdict(list),  # files_by_iteration
        defaultdict(float)  # mpp_sums
    )

def adjust_values(vals):
    return {name: int(val * 1000) for val, name in vals}

def load_circuit_data(circuit_list, directory):
    circuit_data = {}
    for circuit in circuit_list:
        file_path = os.path.join(directory, f'z_3x3_irr_{circuit}.csv')
        circuit_data[circuit] = pd.read_csv(file_path)
    return circuit_data

def get_max_power(df, number):
    row = df[df['Irr'] == number]
    return row.iloc[0]['MPP'] if len(row) == 1 else None

def generate_max_power_dict(adjusted_vals, circuit_dfs):
    max_power_dict = defaultdict(dict)
    for name, number in adjusted_vals.items():
        for circuit, df in circuit_dfs.items():
            max_power = get_max_power(df, number)
            if max_power:
                max_power_dict[circuit][name] = max_power
            else:
                print(f"Warning: No unique match found for {name} with iteration {number} in circuit {circuit}")
    return max_power_dict

def read_csv(file_path):
    try:
        return pd.read_csv(file_path).values.tolist()
    except Exception as e:
        print(f"Error reading {file_path}: {e}")
        return []

def extract_parameters(data, W):
    try:
        V, I, P = zip(*data)
        Voc = next(V[i] for i in range(len(I)) if I[i] < 0.0005)
        Isc, MPP = max(I), max(P)
        MPP_index = P.index(MPP)
        Vmp, Imp = V[MPP_index], I[MPP_index]
        FF = (MPP / (Isc * Voc)) * 100
        Ef = (MPP / W) * 100
        return Voc, Isc, Vmp, Imp, MPP, FF, Ef
    except Exception as e:
        print(f"Error extracting parameters: {e}")
        return None, None, None, None, None, None, None

def perform_calculations(files_by_iteration, max_power_dict, W, mpp_sums, t1, percentage_diffs, percentage_diffs_pos, percentage_diffs_val_mpp):
    for iteration, csv_files in files_by_iteration.items():
        for csv_file in csv_files:
            data = read_csv(csv_file)
            if not data:
                continue
            
            circuit_name = os.path.basename(csv_file).split('_data_iteration_')[0]
            params = extract_parameters(data, W)
            if None in params:
                continue
            
            Voc, Isc, Vmp, Imp, MPP, FF, Ef = params
            reference_mpp = max_power_dict[circuit_name].get(iteration)
            if reference_mpp is None:
                continue
            
            percentage_diff = (MPP - reference_mpp) / reference_mpp
            percentage_diffs[circuit_name].append(percentage_diff)
            mpp_sums[circuit_name] += percentage_diff
            t1[circuit_name].append((MPP, reference_mpp))

            if percentage_diff > 0:
                percentage_diffs_pos[circuit_name].append((percentage_diff, iteration, circuit_name))
            if 0 < percentage_diff < 0.05:
                percentage_diff = 0
            percentage_diffs_val_mpp[circuit_name].append((percentage_diff, iteration, circuit_name))

def create_percentage_diffs_df(percentage_diffs_val_mpp):
    data = [
        {'Circuit': circuit, 'Iteration': iteration, 'PercentageDiff': value}
        for circuit_name, values in percentage_diffs_val_mpp.items()
        for value, iteration, circuit in values
    ]
    return pd.DataFrame(data)

def map_iteration_values(df, vals_dict):
    df['IterationValue'] = df['Iteration'].map(vals_dict)
    return df

def sort_dataframe(df, iteration_order):
    # Ensure 'IterationType' is extracted from 'Iteration'
    df['IterationType'] = df['Iteration'].str.extract(r'([A-Z]+_[A-Z]+)_')
    
    # Add a temporary column for the custom iteration order
    df['IterationOrder'] = df['Iteration'].apply(lambda x: iteration_order.index(x) if x in iteration_order else -1)
    
    # Sort the DataFrame first by IterationValue in descending order, then by the custom IterationOrder
    df = df.sort_values(by=['IterationValue', 'IterationType'], ascending=[False, True])
    
    # Drop the temporary column
    df.drop(columns=['IterationOrder'], inplace=True)
    
    return df

def calculate_medians(df):
    # Ensure there are no missing values in 'PercentageDiff' and 'IterationValue'
    df = df.dropna(subset=['PercentageDiff', 'IterationValue'])

    # Group by 'Circuit' and 'IterationType', then calculate the median and first value
    median_values = df.groupby(['Circuit', 'IterationType']).agg({
        'PercentageDiff': 'median',
        'IterationValue': 'first'
    }).reset_index()

    # Sort by 'IterationValue' in descending order and 'IterationType' in ascending order
    median_values = median_values.sort_values(by=['IterationValue', 'IterationType'], ascending=[False, True])

    return median_values

def find_extremes(df, func):
    indices = df.groupby(['Circuit', 'IterationType'])['PercentageDiff'].apply(func)
    return df.loc[indices].reset_index(drop=True).sort_values(by=['IterationValue', 'IterationType'], ascending=[False, True])

def process_lt_circuit_data(result_df_max, result_df_min, median_values, circ):
    # Filter data for iterations starting with 'Z'
    lt_df_max = result_df_max[result_df_max['Iteration'].str.startswith('Z')]
    lt_df_min = result_df_min[result_df_min['Iteration'].str.startswith('Z')]
    lt_median_values = median_values[median_values['IterationType'].str.startswith('Z')]

    # Filter data for the specific circuit
    lt_circuit_data_max = lt_df_max[lt_df_max['Circuit'] == circ]
    lt_circuit_data_min = lt_df_min[lt_df_min['Circuit'] == circ]
    lt_circuit_data_median = lt_median_values[lt_median_values['Circuit'] == circ]

    # Ensure that you're working with a copy of the DataFrame slice to avoid SettingWithCopyWarning
    lt_circuit_data_max = lt_circuit_data_max.copy()
    lt_circuit_data_min = lt_circuit_data_min.copy()
    lt_circuit_data_median = lt_circuit_data_median.copy()
    
    lt_circuit_data_max.loc[:, 'Cumulative_Percentage_Difference'] = lt_circuit_data_max['PercentageDiff'].cumsum()
    lt_circuit_data_min.loc[:, 'Cumulative_Percentage_Difference'] = lt_circuit_data_min['PercentageDiff'].cumsum()
    lt_circuit_data_median.loc[:, 'Cumulative_Percentage_Difference'] = lt_circuit_data_median['PercentageDiff'].cumsum()
   
    # Return the data in a dictionary
    return {
        'max': lt_circuit_data_max,
        'min': lt_circuit_data_min,
        'median': lt_circuit_data_median
    }

# Execution
t1, percentage_diffs_pos, percentage_diffs_val_mpp, percentage_diffs, files_by_iteration, mpp_sums = initialize_data_structures()

adjusted_vals = adjust_values(Vals)
circuit_dfs = load_circuit_data(['TCT_D', 'TCT', 'SP_D', 'SP'], circuit_irr_directory)
max_power_dict = generate_max_power_dict(adjusted_vals, circuit_dfs)

csv_files = [os.path.join(csv_directory, filename) for filename in os.listdir(csv_directory) if filename.endswith(".csv")]

for file in csv_files:
    iteration = os.path.basename(file).split('_data_iteration_')[1].split('.')[0]
    files_by_iteration[iteration].append(file)

perform_calculations(files_by_iteration, max_power_dict, W, mpp_sums, t1, percentage_diffs, percentage_diffs_pos, percentage_diffs_val_mpp)

percentage_diffs_df = create_percentage_diffs_df(percentage_diffs_val_mpp)

vals_dict = {value: key for key, value in Vals}
percentage_diffs_df = map_iteration_values(percentage_diffs_df, vals_dict)

iteration_order = list(vals_dict.keys())
percentage_diffs_df = sort_dataframe(percentage_diffs_df, iteration_order)

median_values = calculate_medians(percentage_diffs_df)
result_df_max = find_extremes(percentage_diffs_df, 'idxmax')
result_df_min = find_extremes(percentage_diffs_df, 'idxmin')

# Execution
lt_circuit_data = process_lt_circuit_data(result_df_max, result_df_min, median_values, circ)

# Accessing the specific DataFrames:
LT_circuit_data_max = lt_circuit_data['max']
LT_circuit_data_min = lt_circuit_data['min']
LT_circuit_data_median = lt_circuit_data['median']

# -----------------------------------------------------------------------------
# -----------------------------------------------------------------------------
# -----------------------------------------------------------------------------
# SunSimulator results
# -----------------------------------------------------------------------------
# -----------------------------------------------------------------------------
# -----------------------------------------------------------------------------

def suppress_warnings():
    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

def get_column_indices(column_pairs):
    return [(openpyxl.utils.cell.column_index_from_string(c1), openpyxl.utils.cell.column_index_from_string(c2)) for c1, c2 in column_pairs]

def load_excel_files(directory):
    return [f for f in os.listdir(directory) if f.endswith('.xlsx')]

def load_workbook_sheets(file_path, sheet_names):
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheets = {name: workbook[name] for name in sheet_names if name in workbook.sheetnames}
        return sheets
    except Exception as e:
        print(f"Error loading workbook {file_path}: {e}")
        return {}
    
def process_summary_sheet(sheet):
    parameters_irr = {}
    parameters = {}
    comment_value = None
    for row in sheet.iter_rows(min_row=1, max_col=2, values_only=True):
        if row[0] == 'Comment':
            comment_value = row[1]
            break
    
    if comment_value == 'Irr':
        titles = [1000, 889, 778, 667, 556, 444, 333, 222, 111]
        for i, (start_col, end_col) in enumerate(column_indices):
            for row in sheet.iter_rows(min_row=1, min_col=start_col, max_col=end_col, values_only=True):
                parameter = row[0]
                value = row[1]
                if parameter == "Pmpp":
                    parameters_irr[f"Pmpp_{titles[i]}"] = value
                    break
        return {'Comment': comment_value, **parameters_irr}
    
    else:
        start_collecting = False
        for row in sheet.iter_rows(min_row=1, min_col=12, max_col=13, values_only=True):
            parameter = row[0]
            value = row[1]
            if parameter == "Isc":
                start_collecting = True
            if start_collecting:
                parameters[parameter] = value
                if parameter == "ETA":
                    break
        return {'Comment': comment_value, **parameters}

def process_raw_sheet(sheet):
    voltage_data = []
    current_data = []
    found_headers = False
    
    for row in sheet.iter_rows(min_row=1, values_only=True):
        if found_headers:
            voltage = row[6]
            current = row[7]
            if voltage is not None and current is not None:
                voltage_data.append(voltage)
                current_data.append(-current)
            else:
                break
        if row[6] == 'Voltage [V]' and row[7] == 'Current [A]':
            found_headers = True

    return voltage_data, current_data

def calculate_percentage_difference(actual, reference):
    """Calculate percentage difference between actual and reference values."""
    try:
        return ((actual - reference) / reference)
    except ZeroDivisionError:
        return None  # Handle division by zero if reference is zero

def create_comparison_df(df, mapped_df, pmpp_col, iteration_name):
    """Create a DataFrame with Comment, Pmpp, Percentage Difference, and Irradiance."""
    # Create a mapping dictionary from mapped_pmpp_df
    # Includes both Pmpp Value and Irradiance
    pmpp_mapping = mapped_df.set_index('Comment')[['Pmpp Value', 'Irradiance']].to_dict('index')
    
    def get_pmpp_and_irradiance(comment):
        """Retrieve Pmpp Value and Irradiance for a given comment."""
        return pmpp_mapping.get(comment, {'Pmpp Value': None, 'Irradiance': None})

    # Calculate percentage difference and add new columns
    df_comparison = df.copy()
    df_comparison['Irradiance'] = df_comparison['Comment'].apply(lambda c: get_pmpp_and_irradiance(c)['Irradiance'])
    df_comparison['Percentage Difference'] = df_comparison.apply(
        lambda row: calculate_percentage_difference(row[pmpp_col], get_pmpp_and_irradiance(row['Comment'])['Pmpp Value']), axis=1
    )

    # Select only relevant columns
    df_comparison = df_comparison[['Comment', pmpp_col, 'Irradiance', 'Percentage Difference']]
    
    # Rename columns for clarity
    df_comparison.columns = ['Comment', 'Pmpp', 'Irradiance', 'Percentage Difference']

    # Add an iteration column
    df_comparison['Iteration'] = iteration_name
    
    return df_comparison

# Add the color parameter to the plotting function
def plot_cumulative_percentage_difference(ax, df, label, color, mult):
    """Plot cumulative percentage difference with x-axis as sequential integers and specific color."""
    df_sorted = df.sort_values(by='Irradiance', ascending=False).reset_index(drop=True)
    df_sorted['Cumulative Percentage Difference'] = df_sorted['Percentage Difference'].cumsum()
    ax.plot(x, df_sorted['Cumulative Percentage Difference']*mult, linestyle='--', label=label, color=color)
    
    ax.set_xlim(left=-0.5, right=30.5)
    
    ax.set_xlabel('Irradiance', fontsize=16)
    ax.set_ylabel('Cumulative Percentage Difference (%)', fontsize=16)
    ax.set_title(f'Cumulative Percentage Difference vs Irradiance - {circ} {m}', fontsize=18)
    
    ax.set_xticks([1, 29])
    ax.set_xticklabels(['Highest\n(~900)', 'Lowest\n(~100)'], fontsize=13)
    
    ax.tick_params(axis='x', which='both', bottom=False, top=False, labelbottom=True, labelsize=13)
    ax.tick_params(axis='y', which='both', labelsize=13)


irradiance=Irradiance
# def main(directory, irradiance):
suppress_warnings()
column_pairs = [('L', 'M'), ('AJ', 'AK'), ('AV', 'AW'), ('BH', 'BI'), ('BT', 'BU'), ('CF', 'CG'), ('CR', 'CS'), ('DD', 'DE'), ('DP', 'DQ')]
global column_indices
column_indices = get_column_indices(column_pairs)

files = load_excel_files(directory)
results_irr = []
results = []

for file in files:
    file_path = os.path.join(directory, file)
    sheets = load_workbook_sheets(file_path, ['IV-Summary', 'IV-Raw'])
    
    if 'IV-Summary' in sheets:
        summary_sheet = sheets['IV-Summary']
        result = process_summary_sheet(summary_sheet)
        if result.get('Comment') == 'Irr':
            results_irr.append(result)
        else:
            results.append(result)
    
    if 'IV-Raw' in sheets:
        raw_sheet = sheets['IV-Raw']
        process_raw_sheet(raw_sheet)

df_irr = pd.DataFrame(results_irr)
df = pd.DataFrame(results)

filtered_df = df[df['Comment'].apply(lambda x: isinstance(x, str) and bool(re.match(r'^[A-Z]{2}$', x)))]
if 'Pmpp' in filtered_df.columns:
    total_pmpp = filtered_df['Pmpp'].sum()
    print(f"Total Pmpp: {total_pmpp}")
else:
    print("No Pmpp data found.")

if 'Pmpp' in df.columns:
    idxM = df.groupby('Comment')['Pmpp'].idxmax()
    idxm = df.groupby('Comment')['Pmpp'].idxmin()
    highest_pmpp_df = df.loc[idxM].reset_index(drop=True)
    lowest_pmpp_df = df.loc[idxm].reset_index(drop=True)
    median_pmpp_df = df.groupby('Comment').apply(lambda x: x.loc[(x['Pmpp'] - x['Pmpp'].median()).abs().idxmin()]).reset_index(drop=True)

pmpp_values = df_irr.iloc[0].to_dict()
pmpp_values.pop('Comment')
mapped_pmpp_values = [(irr, comment, pmpp_values.get(f'Pmpp_{irr}', 'Not Found')) for irr, comment in irradiance]
mapped_pmpp_df = pd.DataFrame(mapped_pmpp_values, columns=['Irradiance', 'Comment', 'Pmpp Value'])

# Create comparison DataFrames for highest, lowest, and median
highest_comparison_df = create_comparison_df(highest_pmpp_df, mapped_pmpp_df, 'Pmpp', 'Highest')
lowest_comparison_df = create_comparison_df(lowest_pmpp_df, mapped_pmpp_df, 'Pmpp', 'Lowest')
median_comparison_df = create_comparison_df(median_pmpp_df, mapped_pmpp_df, 'Pmpp', 'Median')

# Combine all comparison DataFrames into a single DataFrame
combined_comparison_df = pd.concat([highest_comparison_df, lowest_comparison_df, median_comparison_df], ignore_index=True)

# Reorder the DataFrame by 'Irradiance' (highest to lowest) and 'Comment' (alphabetically)
sorted_combined_comparison_df = combined_comparison_df.sort_values(by=['Irradiance', 'Comment'], ascending=[False, True])

# Filter data by Iteration
df_highest = sorted_combined_comparison_df[sorted_combined_comparison_df['Iteration'] == 'Highest']
df_lowest = sorted_combined_comparison_df[sorted_combined_comparison_df['Iteration'] == 'Lowest']
df_median = sorted_combined_comparison_df[sorted_combined_comparison_df['Iteration'] == 'Median']

# Create a single plot
plt.figure(figsize=(14, 7))

# Plot cumulative percentage difference for each iteration
plot_cumulative_percentage_difference(plt.gca(), df_highest, 'Highest IRL', 'pink', 0.7)
plot_cumulative_percentage_difference(plt.gca(), df_lowest, 'Lowest IRL', 'lightblue', 1.3)
plot_cumulative_percentage_difference(plt.gca(), df_median, 'Median IRL', 'gray', 1)

# Ensure that x values and DataFrames for additional lines are properly defined
plt.plot(x, LT_circuit_data_max['Cumulative_Percentage_Difference']*0.7, label='Highest LTS', color='red')
plt.plot(x, LT_circuit_data_min['Cumulative_Percentage_Difference']*1.3, label='Lowest LTS', color='blue')
plt.plot(x, LT_circuit_data_median['Cumulative_Percentage_Difference']*1, label='Median LTS', color='black')

# Add vertical lines at specific x-coordinates
x_lines = [2, 7, 11, 19, 24, 27, 29]  # Example x-coordinates for vertical lines
x_ls = [0,30]
for x_line in x_lines:
    plt.axvline(x=x_line, color='green', linestyle='--', linewidth=1, alpha=0.5)

for x_l in x_ls:
    plt.axvline(x=x_l, color='black', linestyle='--', linewidth=1)

plt.legend(fontsize=13)

# Show the plot
plt.show()
    
# main(directory, Irradiance)
