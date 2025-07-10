# -*- coding: utf-8 -*-
"""
Created on Thu Sep 19 14:50:08 2024

@author: ruiui
"""

import openpyxl
import warnings
import os
import numpy as np
import pandas as pd
import re

order = ['AA', 'AI', 'AP', 
         'AB', 'AG', 'AJ', 'AS', 'AX', 
         'AC', 'AD', 'AH', 'AK', 
         'AL', 'AN', 'AQ', 'AV', 'AY', 'BA', 'BD', 'BE', 
         'AE', 'AM', 'AT', 'AZ', 'BB', 
         'AF', 'AW', 'BC', 
         'AO', 'AU', 
         'AR']

Irradiance = [(1000, 'AA'), (1000, 'AB'), (1000, 'AC'), (1000, 'AD'), (1000, 'AE'), 
              (1000, 'AF'), (1000, 'AG'), (1000, 'AH'), (1000, 'AI'), (1000, 'AJ'), 
              (1000, 'AK'), (1000, 'AL'), (1000, 'AM'), (1000, 'AN'), (1000, 'AO'), 
              (1000, 'AP'), (1000, 'AQ'), (1000, 'AR'), (1000, 'AS'), (1000, 'AT'), 
              (1000, 'AU'), (1000, 'AV'), (1000, 'AW'), (1000, 'AX'), (1000, 'AY'), 
              (1000, 'AZ'), (1000, 'BA'), (1000, 'BB'), (1000, 'BC'), (1000, 'BD'), 
              (1000, 'BE')]

Vals = ['Z_AA_0','Z_AA_1','Z_AA_2','Z_AA_3', 
        'Z_AB_0','Z_AB_1','Z_AB_2','Z_AB_3', 
        'Z_AC_0','Z_AC_1','Z_AC_2','Z_AC_3', 
        'Z_AD_0','Z_AD_1','Z_AD_2','Z_AD_3', 
        'Z_AE_0','Z_AE_1','Z_AE_2','Z_AE_3', 
        'Z_AF_0','Z_AF_1','Z_AF_2','Z_AF_3', 
        'Z_AG_0','Z_AG_1','Z_AG_2','Z_AG_3', 
        'Z_AH_0','Z_AH_1',
        'Z_AI_0','Z_AI_1','Z_AI_2','Z_AI_3', 
        'Z_AJ_0','Z_AJ_1','Z_AJ_2','Z_AJ_3', 
        'Z_AK_0','Z_AK_1', 
        'Z_AL_0','Z_AL_1','Z_AL_2','Z_AL_3', 
        'Z_AM_0',
        'Z_AN_0','Z_AN_1','Z_AN_2','Z_AN_3',
        'Z_AO_0','Z_AO_1','Z_AO_2','Z_AO_3',
        'Z_AP_0', 
        'Z_AQ_0', 
        'Z_AR_0', 
        'Z_AS_0','Z_AS_1','Z_AS_2','Z_AS_3', 
        'Z_AT_0','Z_AT_1','Z_AT_2','Z_AT_3', 
        'Z_AU_0','Z_AU_1','Z_AU_2','Z_AU_3', 
        'Z_AV_0','Z_AV_1','Z_AV_2','Z_AV_3', 
        'Z_AW_0','Z_AW_1',
        'Z_AX_0','Z_AX_1', 
        'Z_AY_0','Z_AY_1', 
        'Z_AZ_0','Z_AZ_1', 
        'Z_BA_0','Z_BA_1','Z_BA_2','Z_BA_3', 
        'Z_BB_0','Z_BB_1','Z_BB_2','Z_BB_3', 
        'Z_BC_0','Z_BC_1','Z_BC_2','Z_BC_3', 
        'Z_BD_0','Z_BD_1','Z_BD_2','Z_BD_3', 
        'Z_BE_0','Z_BE_1','Z_BE_2','Z_BE_3'
        ]

# -----------------------------------------------------------------------------

circ = 'TCT'

circ = 'TCT_D'

circ = 'SP'

circ = 'SP_D'

circs = ['TCT', 'TCT_D', 'SP', 'SP_D']

reference_value = 22.5089475498911

r_sp_d =  22.445806027077
r_sp =    22.413338691756
r_tct_d = 22.394378963305
r_tct =   22.508947549891
# -----------------------------------------------------------------------------

base_directory = 'C:/Users/ruiui/Desktop/solar plug and play old stuff/sun simulator new/'

# Initialize the final DataFrame
final_pmpp_df = None

for circ in circs:
    directory = f'{base_directory}{circ}/'
    
    
    def suppress_warnings():
        warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
    
    def get_column_indices(column_pairs):
        return [(openpyxl.utils.cell.column_index_from_string(c1), openpyxl.utils.cell.column_index_from_string(c2)) for c1, c2 in column_pairs]
    
    def load_excel_files(directory):
        return [f for f in os.listdir(directory) if f.endswith('.xlsx')]
    
    def load_workbook_sheets(file_path, sheet_names):
        try:
            workbook = openpyxl.load_workbook(file_path)
            sheets = {name: workbook[name] for name in sheet_names if name in workbook.sheetnames}
            return sheets
        except Exception as e:
            print(f"Error loading workbook {file_path}: {e}")
            return {}
        
    def process_summary_sheet(sheet, filename):
        parameters_irr = {}
        parameters = {}
        comment_value = None
        for row in sheet.iter_rows(min_row=1, max_col=2, values_only=True):
            if row[0] == 'Comment':
                comment_value = row[1]
                break
        
        if comment_value == 'Irr':
            titles = [1000, 889, 778, 667, 556, 444, 333, 222, 111]
            for i, (start_col, end_col) in enumerate(column_indices):
                for row in sheet.iter_rows(min_row=1, min_col=start_col, max_col=end_col, values_only=True):
                    parameter = row[0]
                    value = row[1]
                    if parameter == "Pmpp":
                        parameters_irr[f"Pmpp_{titles[i]}"] = value
                        break
            return {'Comment': comment_value, 'Filename': filename, **parameters_irr}
        
        else:
            start_collecting = False
            for row in sheet.iter_rows(min_row=1, min_col=12, max_col=13, values_only=True):
                parameter = row[0]
                value = row[1]
                if parameter == "Isc":
                    start_collecting = True
                if start_collecting:
                    parameters[parameter] = value
                    if parameter == "ETA":
                        break
            return {'Comment': comment_value, 'Filename': filename, **parameters}
    
    def calculate_percentage_difference(actual, reference):
        """Calculate percentage difference between actual and reference values."""
        try:
            return ((actual - reference) / reference)
        except ZeroDivisionError:
            return None  # Handle division by zero if reference is zero
    
    def create_comparison_df(df, mapped_df, pmpp_col, iteration_name):
        """Create a DataFrame with Comment, Pmpp, Percentage Difference, and Irradiance."""
        # Create a mapping dictionary from mapped_pmpp_df
        # Includes both Pmpp Value and Irradiance
        pmpp_mapping = mapped_df.set_index('Comment')[['Pmpp Value', 'Irradiance']].to_dict('index')
        
        def get_pmpp_and_irradiance(comment):
            """Retrieve Pmpp Value and Irradiance for a given comment."""
            return pmpp_mapping.get(comment, {'Pmpp Value': None, 'Irradiance': None})
    
        # Calculate percentage difference and add new columns
        df_comparison = df.copy()
        df_comparison['Irradiance'] = df_comparison['Comment'].apply(lambda c: get_pmpp_and_irradiance(c)['Irradiance'])
        df_comparison['Percentage Difference'] = df_comparison.apply(lambda row: calculate_percentage_difference(row[pmpp_col], get_pmpp_and_irradiance(row['Comment'])['Pmpp Value']), axis=1
        )
    
        # Select only relevant columns
        df_comparison = df_comparison[['Filename', 'Comment', pmpp_col, 'Irradiance', 'Percentage Difference']]
        
        # Rename columns for clarity
        df_comparison.columns = ['Filename', 'Comment', 'Pmpp', 'Irradiance', 'Percentage Difference']
    
        # Add an iteration column
        df_comparison['Iteration'] = iteration_name
        
        return df_comparison
    
    def create_comparison_df_T(df, mapped_df, pmpp_col):
        """Create a DataFrame with Comment, Pmpp, Percentage Difference, and Irradiance."""
        # Create a mapping dictionary from mapped_df
        # Includes both Pmpp Value and Irradiance
        pmpp_mapping = mapped_df.set_index('Comment')[['Pmpp Value', 'Irradiance']].to_dict('index')
        
        def get_pmpp_and_irradiance_T(comment):
            """Retrieve Pmpp Value and Irradiance for a given comment."""
            return pmpp_mapping.get(comment, {'Pmpp Value': None, 'Irradiance': None})
    
        # Calculate percentage difference and add new columns
        df_comparison = df.copy()
        df_comparison['Irradiance'] = df_comparison['Comment'].apply(lambda c: get_pmpp_and_irradiance_T(c)['Irradiance'])
        df_comparison['Percentage Difference'] = df_comparison.apply(lambda row: calculate_percentage_difference(row[pmpp_col], get_pmpp_and_irradiance_T(row['Comment'])['Pmpp Value']), axis=1)
    
        # Select only relevant columns
        df_comparison = df_comparison[['Filename', 'Comment', pmpp_col, 'Irradiance', 'Percentage Difference']]
        
        # Rename columns for clarity
        df_comparison.columns = ['Filename', 'Comment', 'Pmpp', 'Irradiance', 'Percentage Difference']
    
        return df_comparison
    
    def extract_iv_data_from_file_ss(sheet, filename):
        voltage_data = []
        current_data = []
        found_headers = False
        
        for row in sheet.iter_rows(min_row=1, values_only=True):
            if found_headers:
                voltage = row[6]
                current = row[7]
                if voltage is not None and current is not None:
                    voltage_data.append(voltage)
                    current_data.append(-current)
                else:
                    break
            if row[6] == 'Voltage [V]' and row[7] == 'Current [A]':
                found_headers = True
    
        return voltage_data, current_data
    
    
    column_pairs = [('L', 'M'), ('AJ', 'AK'), ('AV', 'AW'), ('BH', 'BI'), ('BT', 'BU'), ('CF', 'CG'), ('CR', 'CS'), ('DD', 'DE'), ('DP', 'DQ')]
    global column_indices
    column_indices = get_column_indices(column_pairs)
    
    files = load_excel_files(directory)
    results_irr = []
    results = []
    
    for file in files:
        file_path = os.path.join(directory, file)
        sheets = load_workbook_sheets(file_path, ['IV-Summary', 'IV-Raw'])
        
        if 'IV-Summary' in sheets:
            summary_sheet = sheets['IV-Summary']
            raw_sheet = sheets['IV-Raw']
            result = process_summary_sheet(summary_sheet, file)
            if result.get('Comment') == 'Irr':
                results_irr.append(result)
            else:
                results.append(result)

    df_irr = pd.DataFrame(results_irr)
    df = pd.DataFrame(results)

    # Drop unnecessary columns and rename the Pmpp column
    a_df_sunsim = df.drop(columns=['Filename', 'Isc', 'Voc', 'FF', 'Rs', 'Vmpp', 'Impp', 'ETA'], inplace=False)
    a_df_sunsim = a_df_sunsim.rename(columns={'Comment': 'base_code', 'Pmpp': circ}, inplace=False)

    # Merge or update final_pmpp_df
    if final_pmpp_df is None:
        final_pmpp_df = a_df_sunsim  # Initialize with the first DataFrame
    else:
        # Instead of merging, use concat to add new column for the current circ
        final_pmpp_df = pd.concat([final_pmpp_df, a_df_sunsim[circ]], axis=1)

# Now final_pmpp_df should have the desired shape
final_pmpp_df['it'] = Vals  # Add your 'it' column

# Select and reorder columns
final_pmpp_df = final_pmpp_df[['it', 'base_code', 'SP', 'SP_D', 'TCT', 'TCT_D']]

# Rename and reorder columns
final_pmpp_df = final_pmpp_df.rename(columns={
    'it': 'it', 
    'base_code': 'base_code', 
    'SP': 'SP', 
    'SP_D': 'SP_Dio', 
    'TCT': 'TCT', 
    'TCT_D': 'TCT_Dio'
})[['it', 'base_code', 'SP', 'SP_Dio', 'TCT', 'TCT_Dio']]

# Save the final merged DataFrame to a CSV file
output_file_path = "C:/Users/ruiui/Desktop/iteration data/tabelas_pd_cumu/sun_simulator_3x3_results.csv"
final_pmpp_df.to_csv(output_file_path, index=False)

# Optionally, print a message confirming the save
print(f"DataFrame saved to {output_file_path}")











