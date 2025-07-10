# -*- coding: utf-8 -*-
"""
Created on Fri Oct 11 16:48:39 2024

@author: ruiui
"""

import os
import openpyxl
import warnings
import matplotlib.pyplot as plt
import pandas as pd
import numpy as np
from scipy.interpolate import interp1d

# Suppress the specific warning from openpyxl
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# Function to extract data from .xlsx files (IV data and comments)
def read_xlsx_files(directory):
    files = [f for f in os.listdir(directory) if f.endswith('.xlsx')]
    comment_data = {}
    
    for file in files:
        file_path = os.path.join(directory, file)
        workbook = openpyxl.load_workbook(file_path)
        sheeti = workbook['IV-Summary']
        
        # Extract the comment
        comment_value = None
        for row in sheeti.iter_rows(min_row=1, max_col=2, values_only=True):
            if row[0] == 'Comment':
                comment_value = row[1]
                break
        
        sheet = workbook['IV-Raw']
        voltage_data = []
        current_data = []
        found_headers = False
        
        for row in sheet.iter_rows(min_row=1, values_only=True):
            if found_headers:
                voltage = row[6]
                current = row[7]
                if voltage is not None and current is not None:
                    voltage_data.append(voltage)
                    current_data.append(-current)  # Assuming negative currents for IV curve
                else:
                    break
            if row[6] == 'Voltage [V]' and row[7] == 'Current [A]':
                found_headers = True
        
        if comment_value not in comment_data:
            comment_data[comment_value] = []
        comment_data[comment_value].append((voltage_data, current_data))
    
    return comment_data

# Updated Function to calculate solar cell parameters (Voc, Isc, Vmp, Imp, Pmax, FF, Ef)
def calculate_parameters(voltage, current):
    # Create a DataFrame for voltage and current
    df = pd.DataFrame({'Voltage': voltage, 'Current': current})
    
    # Filter the data to keep only positive voltage and current
    df_filtered = df[(df['Voltage'] > 0) & (df['Current'] > 0)]
    
    # Extract the filtered Voltage and Current as numpy arrays
    V = df_filtered['Voltage'].values
    I = df_filtered['Current'].values
    P = V * I  # Power calculation
    
    # Interpolate the data to have the same voltage points for all curves
    interp_func = interp1d(V, I, kind='linear', fill_value="extrapolate")
    voltage_axis = np.linspace(V.min(), V.max(), num=100)  # Create a uniform voltage axis
    interpolated_current = interp_func(voltage_axis)

    interp_func = interp1d(V, P, kind='linear', fill_value="extrapolate")
    interpolated_power = interp_func(voltage_axis)
    
    # Create a new DataFrame for the interpolated data
    df_interpolated = pd.DataFrame({
        'Voltage (V)': voltage_axis,
        'Current (A)': interpolated_current,
        'Power (W)': interpolated_power
    })
    
    # Calculate key parameters
    Voc = np.max(df_interpolated['Voltage (V)'])  # Open-circuit voltage
    Isc = np.max(df_interpolated['Current (A)'])  # Short-circuit current
    Pmax = np.max(df_interpolated['Power (W)'])  # Maximum power point
    Vmp = df_interpolated['Voltage (V)'].iloc[np.argmax(df_interpolated['Power (W)'])]  # Voltage at max power
    Imp = df_interpolated['Current (A)'].iloc[np.argmax(df_interpolated['Power (W)'])]  # Current at max power

    # Fill Factor (FF)
    FF = (Pmax / (Isc * Voc))
    
    # Efficiency (Ef) - Placeholder for W (irradiance or area factor)
    W = 12.5 * 12.5 * 0.1  # You can adjust W based on your system or dataset
    Ef = (Pmax / W)
    
    return Voc, Isc, Vmp, Imp, Pmax, FF, Ef

# Function to calculate parameters for all curves under each comment and compute variability
def process_comment_curves(comment_data):
    comment_variability = {}

    for comment, curves in comment_data.items():
        parameters_list = []
        
        for voltage_data, current_data in curves:
            # Calculate parameters for each curve
            Voc, Isc, Vmp, Imp, Pmax, FF, Ef = calculate_parameters(voltage_data, current_data)
            parameters_list.append([Voc, Isc, Vmp, Imp, Pmax, FF, Ef])
        
        # Convert the list to a DataFrame for easier analysis
        df_params = pd.DataFrame(parameters_list, columns=['Voc', 'Isc', 'Vmp', 'Imp', 'MPP', 'FF', 'Efficiency'])
        
        # Calculate mean and standard deviation for each parameter
        param_means = df_params.mean()
        param_stds = df_params.std()
        
        # Store the results for the current comment
        comment_variability[comment] = {
            'mean': param_means,
            'std': param_stds,
            'individual': df_params  # Store the individual parameter values
        }
    
    return comment_variability

# Function to save standard deviations to a CSV file
def save_std_to_csv(variability_results, output_file):
    # Create a DataFrame for storing the results
    std_df = pd.DataFrame(columns=['File', 'Voc', 'Isc', 'Vmp', 'Imp', 'MPP', 'FF', 'Efficiency'])
    
    for comment, stats in variability_results.items():
        std_row = [comment] + stats['std'].tolist()  # Combine comment and std values
        std_df.loc[len(std_df)] = std_row  # Append to DataFrame
    
    # Save the DataFrame to a CSV file
    std_df.to_csv(output_file, index=False)
    
# Optional: Function to plot IV curves
def plot_iv_curves(comment_data):
    for i, (comment, data_list) in enumerate(comment_data.items()):
        plt.figure(figsize=(10, 6))

        all_voltages = []
        all_currents = []

        for voltage_data, current_data in data_list:
            plt.plot(voltage_data, current_data, label=f"File {data_list.index((voltage_data, current_data)) + 1}", linewidth=0.5)
            all_voltages.extend(voltage_data)
            all_currents.extend(current_data)

        # Interpolate the curve
        min_voltage = min(all_voltages) + 0.1
        max_voltage = max(all_voltages) - 0.1
        voltage_curve = np.linspace(min_voltage, max_voltage, 1000)
        current_curve = np.interp(voltage_curve, all_voltages, all_currents, left=0, right=0)

        # Plot the fitted curve
        plt.plot(voltage_curve, current_curve, color='black', label='Fitted Curve', linewidth=0.5)

        plt.title(f"IV Curves and Fitted Curve for {comment}")
        plt.xlabel("Voltage [V]")
        plt.ylabel("Current [A]")
        plt.ylim(0, 1)
        plt.xlim(0, 4)
        plt.legend()
        plt.grid(True)
        plt.show()

# Directory containing the files
directory = "C:/Users/ruiui/Desktop/solar plug and play old stuff/sun simulator new/cel_new_2024"
output_file = "C:/Users/ruiui/Desktop/solar plug and play old stuff/sun_simulator_variability_results_batch1.csv"

# directory = "C:/Users/ruiui/Desktop/sun simulator new/cel_new_2024/batch2"
# output_file = "C:/Users/ruiui/Desktop/sun_simulator_variability_results_batch2.csv"

# directory = "C:/Users/ruiui/Desktop/sun simulator new/cel_new_2024/batch3"
# output_file = "C:/Users/ruiui/Desktop/sun_simulator_variability_results_batch3.csv"


# Call functions
comment_data = read_xlsx_files(directory)
plot_iv_curves(comment_data)  # Optional: To visualize the IV curves

# Function to print the results in a readable format
def print_variability_results(variability_results):
    for comment, stats in variability_results.items():
        print(f"Comment: {comment}")
        print("Means:")
        print(stats['mean'])
        print("Standard Deviations:")
        print(stats['std'])
        print("=" * 40)

# Call functions
variability_results = process_comment_curves(comment_data)

# Print the results
print_variability_results(variability_results)

# Save standard deviation results to a CSV file
save_std_to_csv(variability_results, output_file)

# Output the variability results
for comment, stats in variability_results.items():
    print(f"Comment: {comment}")
    # print("Mean values:")
    # print(stats['mean'])
    print("Standard deviation values:")
    print(stats['std'])
    print()